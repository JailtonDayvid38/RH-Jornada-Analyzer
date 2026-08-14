import streamlit as st
import pandas as pd
from io import BytesIO

from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# =========================================================
# CONFIGURAÇÃO DA PÁGINA
# =========================================================

st.set_page_config(
    page_title="RH Jornada Analyzer",
    page_icon="🕒",
    layout="wide"
)

st.title("🕒 RH Jornada Analyzer")

st.write(
    """
    Sistema para análise de jornada de trabalho,
    interjornada e possíveis inconsistências.
    """
)

st.divider()


# =========================================================
# FUNÇÕES AUXILIARES
# =========================================================

def decimal_para_hhmm(horas):
    """
    Converte horas decimais para o formato HH:MM.

    Exemplo:
    10.5 -> 10:30
    0.5  -> 00:30
    """

    if pd.isna(horas):
        return "-"

    minutos_totais = round(float(horas) * 60)

    horas_inteiras = minutos_totais // 60
    minutos = minutos_totais % 60

    return f"{horas_inteiras:02d}:{minutos:02d}"


def gerar_excel(resultado_completo, ocorrencias, resumo_colaborador):
    """
    Gera um relatório Excel em memória.

    O arquivo possui três abas:
    - Resultado Completo
    - Ocorrências
    - Resumo por Colaborador
    """

    buffer = BytesIO()

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl"
    ) as writer:

        resultado_completo.to_excel(
            writer,
            sheet_name="Resultado Completo",
            index=False
        )

        ocorrencias.to_excel(
            writer,
            sheet_name="Ocorrências",
            index=False
        )

        resumo_colaborador.to_excel(
            writer,
            sheet_name="Resumo Colaborador",
            index=False
        )

        # -------------------------------------------------
        # FORMATAÇÃO DAS ABAS
        # -------------------------------------------------

        for nome_aba in writer.book.sheetnames:

            planilha = writer.book[nome_aba]

            planilha.freeze_panes = "A2"
            planilha.auto_filter.ref = planilha.dimensions

            # Cabeçalho
            for celula in planilha[1]:

                celula.font = Font(
                    bold=True
                )

                celula.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # Ajuste automático da largura das colunas
            for coluna in planilha.columns:

                maior_tamanho = 0

                letra_coluna = get_column_letter(
                    coluna[0].column
                )

                for celula in coluna:

                    if celula.value is not None:

                        tamanho = len(
                            str(celula.value)
                        )

                        if tamanho > maior_tamanho:
                            maior_tamanho = tamanho

                planilha.column_dimensions[
                    letra_coluna
                ].width = min(
                    maior_tamanho + 2,
                    45
                )

    buffer.seek(0)

    return buffer


# =========================================================
# IMPORTAÇÃO
# =========================================================

st.subheader("1. Importar arquivo de ponto")

arquivo = st.file_uploader(
    "Selecione uma planilha Excel ou arquivo CSV",
    type=["xlsx", "csv"]
)


if arquivo is not None:

    # =====================================================
    # LEITURA
    # =====================================================

    try:

        if arquivo.name.lower().endswith(".xlsx"):

            dados = pd.read_excel(
                arquivo
            )

        else:

            dados = pd.read_csv(
                arquivo
            )

        st.success(
            "Arquivo carregado com sucesso!"
        )

    except Exception as erro:

        st.error(
            f"Erro ao ler o arquivo: {erro}"
        )

        st.stop()


    # =====================================================
    # VALIDAÇÃO DAS COLUNAS
    # =====================================================

    colunas_obrigatorias = [
        "Matrícula",
        "Nome",
        "Data",
        "Entrada",
        "Saída"
    ]

    colunas_faltantes = [
        coluna
        for coluna in colunas_obrigatorias
        if coluna not in dados.columns
    ]

    if colunas_faltantes:

        st.error(
            "A planilha não possui todas "
            "as colunas obrigatórias."
        )

        st.write(
            "Colunas ausentes:",
            colunas_faltantes
        )

        st.stop()


    # =====================================================
    # TRATAMENTO DOS DADOS
    # =====================================================

    dados["Data"] = pd.to_datetime(
        dados["Data"],
        errors="coerce"
    )

    dados["DataHoraEntrada"] = pd.to_datetime(
        dados["Data"].dt.strftime("%Y-%m-%d")
        + " "
        + dados["Entrada"].astype(str),
        errors="coerce"
    )

    dados["DataHoraSaida"] = pd.to_datetime(
        dados["Data"].dt.strftime("%Y-%m-%d")
        + " "
        + dados["Saída"].astype(str),
        errors="coerce"
    )


    # =====================================================
    # JORNADA APÓS A MEIA-NOITE
    # =====================================================

    # Exemplo:
    # Entrada 22:00
    # Saída   06:00
    #
    # A saída pertence ao dia seguinte.

    jornada_noturna = (
        dados["DataHoraSaida"]
        < dados["DataHoraEntrada"]
    )

    dados.loc[
        jornada_noturna,
        "DataHoraSaida"
    ] = (
        dados.loc[
            jornada_noturna,
            "DataHoraSaida"
        ]
        + pd.Timedelta(days=1)
    )


    # =====================================================
    # ORDENAÇÃO
    # =====================================================

    dados = dados.sort_values(
        [
            "Matrícula",
            "DataHoraEntrada"
        ]
    ).reset_index(drop=True)


    # =====================================================
    # SAÍDA ANTERIOR
    # =====================================================

    dados["SaidaAnterior"] = (
        dados
        .groupby("Matrícula")[
            "DataHoraSaida"
        ]
        .shift(1)
    )


    # =====================================================
    # CÁLCULO DA INTERJORNADA
    # =====================================================

    dados["InterjornadaHoras"] = (
        (
            dados["DataHoraEntrada"]
            - dados["SaidaAnterior"]
        )
        .dt.total_seconds()
        / 3600
    )


    # =====================================================
    # CONFIGURAÇÕES DA ANÁLISE
    # =====================================================

    st.sidebar.header(
        "⚙️ Configurações"
    )

    limite_interjornada = (
        st.sidebar.number_input(
            "Limite mínimo entre jornadas (horas)",
            min_value=1.0,
            max_value=24.0,
            value=11.0,
            step=0.5
        )
    )


    # =====================================================
    # DÉFICIT
    # =====================================================

    dados["DeficitHoras"] = (
        limite_interjornada
        - dados["InterjornadaHoras"]
    ).clip(lower=0)


    # =====================================================
    # STATUS
    # =====================================================

    def classificar_interjornada(horas):

        if pd.isna(horas):

            return "Primeira jornada"

        if horas < limite_interjornada:

            return "⚠️ Interjornada inferior ao limite"

        return "✅ Regular"


    dados["Status"] = (
        dados["InterjornadaHoras"]
        .apply(classificar_interjornada)
    )


    # =====================================================
    # FILTROS
    # =====================================================

    st.sidebar.header(
        "🔎 Filtros"
    )

    nomes = sorted(
        dados["Nome"]
        .dropna()
        .unique()
        .tolist()
    )

    colaborador = (
        st.sidebar.selectbox(
            "Colaborador",
            ["Todos"] + nomes
        )
    )


    data_minima = (
        dados["DataHoraEntrada"]
        .min()
        .date()
    )

    data_maxima = (
        dados["DataHoraEntrada"]
        .max()
        .date()
    )


    periodo = (
        st.sidebar.date_input(
            "Período",
            value=(
                data_minima,
                data_maxima
            ),
            min_value=data_minima,
            max_value=data_maxima
        )
    )


    # =====================================================
    # APLICAÇÃO DOS FILTROS
    # =====================================================

    dados_filtrados = dados.copy()


    if colaborador != "Todos":

        dados_filtrados = (
            dados_filtrados[
                dados_filtrados["Nome"]
                == colaborador
            ]
        )


    if isinstance(
        periodo,
        (tuple, list)
    ) and len(periodo) == 2:

        inicio = pd.Timestamp(
            periodo[0]
        )

        fim = (
            pd.Timestamp(
                periodo[1]
            )
            + pd.Timedelta(days=1)
            - pd.Timedelta(seconds=1)
        )

        dados_filtrados = (
            dados_filtrados[
                (
                    dados_filtrados[
                        "DataHoraEntrada"
                    ] >= inicio
                )
                &
                (
                    dados_filtrados[
                        "DataHoraEntrada"
                    ] <= fim
                )
            ]
        )


    # =====================================================
    # FORMATAÇÃO HH:MM
    # =====================================================

    dados_filtrados[
        "Interjornada"
    ] = (
        dados_filtrados[
            "InterjornadaHoras"
        ]
        .apply(decimal_para_hhmm)
    )


    dados_filtrados[
        "Déficit"
    ] = (
        dados_filtrados[
            "DeficitHoras"
        ]
        .apply(decimal_para_hhmm)
    )


    # =====================================================
    # INDICADORES
    # =====================================================

    total_colaboradores = (
        dados_filtrados[
            "Matrícula"
        ]
        .nunique()
    )

    total_registros = (
        len(
            dados_filtrados
        )
    )

    total_ocorrencias = (
        dados_filtrados[
            "Status"
        ]
        .eq(
            "⚠️ Interjornada inferior ao limite"
        )
        .sum()
    )

    total_deficit = (
        dados_filtrados[
            "DeficitHoras"
        ]
        .sum()
    )


    # =====================================================
    # DASHBOARD
    # =====================================================

    st.subheader(
        "📊 Resumo da análise"
    )

    col1, col2, col3, col4 = (
        st.columns(4)
    )


    col1.metric(
        "Colaboradores",
        total_colaboradores
    )

    col2.metric(
        "Registros analisados",
        total_registros
    )

    col3.metric(
        "Ocorrências",
        int(total_ocorrencias)
    )

    col4.metric(
        "Déficit total",
        decimal_para_hhmm(
            total_deficit
        )
    )


    st.caption(
        f"Parâmetro atual de interjornada: "
        f"{limite_interjornada:g} horas."
    )

    st.divider()


    # =====================================================
    # RESULTADO COMPLETO
    # =====================================================

    st.subheader(
        "2. Resultado da análise"
    )


    resultado = (
        dados_filtrados[
            [
                "Matrícula",
                "Nome",
                "DataHoraEntrada",
                "DataHoraSaida",
                "SaidaAnterior",
                "Interjornada",
                "Déficit",
                "Status"
            ]
        ]
        .copy()
    )


    resultado.columns = [
        "Matrícula",
        "Nome",
        "Entrada",
        "Saída",
        "Saída Anterior",
        "Interjornada",
        "Déficit",
        "Status"
    ]


    st.dataframe(
        resultado,
        use_container_width=True,
        hide_index=True
    )


    # =====================================================
    # OCORRÊNCIAS
    # =====================================================

    ocorrencias = (
        resultado[
            resultado["Status"]
            == (
                "⚠️ Interjornada "
                "inferior ao limite"
            )
        ]
        .copy()
    )


    st.subheader(
        "3. Ocorrências identificadas"
    )


    if ocorrencias.empty:

        st.success(
            "Nenhuma ocorrência encontrada "
            "para os filtros selecionados."
        )

    else:

        st.warning(
            f"Foram encontradas "
            f"{len(ocorrencias)} ocorrência(s)."
        )

        st.dataframe(
            ocorrencias,
            use_container_width=True,
            hide_index=True
        )


    # =====================================================
    # RESUMO POR COLABORADOR
    # =====================================================

    st.subheader(
        "4. Resumo por colaborador"
    )


    resumo_base = (
        dados_filtrados
        .groupby(
            [
                "Matrícula",
                "Nome"
            ],
            as_index=False
        )
        .agg(
            Registros=(
                "Nome",
                "size"
            ),

            Ocorrencias=(
                "Status",
                lambda x: (
                    x
                    == "⚠️ Interjornada inferior ao limite"
                ).sum()
            ),

            DeficitHoras=(
                "DeficitHoras",
                "sum"
            )
        )
    )


    resumo_base[
        "Déficit Total"
    ] = (
        resumo_base[
            "DeficitHoras"
        ]
        .apply(decimal_para_hhmm)
    )


    resumo_colaborador = (
        resumo_base[
            [
                "Matrícula",
                "Nome",
                "Registros",
                "Ocorrencias",
                "Déficit Total"
            ]
        ]
        .copy()
    )


    resumo_colaborador.rename(
        columns={
            "Ocorrencias": "Ocorrências"
        },
        inplace=True
    )


    st.dataframe(
        resumo_colaborador,
        use_container_width=True,
        hide_index=True
    )


    # =====================================================
    # EXPORTAÇÃO
    # =====================================================

    st.subheader(
        "5. Exportar relatório"
    )


    arquivo_excel = gerar_excel(
        resultado,
        ocorrencias,
        resumo_colaborador
    )


    st.download_button(
        label="📥 Baixar relatório em Excel",
        data=arquivo_excel,
        file_name=(
            "relatorio_interjornada.xlsx"
        ),
        mime=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


    # =====================================================
    # OBSERVAÇÃO
    # =====================================================

    st.info(
        """
        O sistema identifica intervalos inferiores ao
        parâmetro configurado entre jornadas.

        O déficit representa apenas a diferença entre
        o intervalo encontrado e o limite configurado.

        O resultado não deve ser interpretado
        automaticamente como quantidade de horas a pagar.
        """
    )